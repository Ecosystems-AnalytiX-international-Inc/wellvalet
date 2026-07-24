"""Generate an Excel checklist for launching WellValet on the App Store & Play Store.

Sheets:
  1. Master Checklist    - all cross-store + store-specific tasks
  2. Apple App Store     - iOS-only requirements
  3. Google Play         - Android-only requirements
  4. Assets Reference    - required image sizes & formats
  5. Legal & Compliance  - privacy policy / data safety / ratings
  6. Summary             - counts by status, category and owner
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUT = Path(__file__).resolve().parent.parent / "STORE_LAUNCH_CHECKLIST.xlsx"


# ─── Palette ────────────────────────────────────────────────────────────────
GREEN_DARK = "2D6A2D"
GREEN_MID = "42D674"
GREEN_LIGHT = "E3F0A3"
GREY_HEAD = "1F2937"
GREY_ALT = "F3F4F6"
BLUE = "1565C0"
ORANGE = "F59E0B"
RED = "DC2626"

thin = Side(style="thin", color="D0D0D0")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def header_fill(color=GREEN_DARK):
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


def alt_fill():
    return PatternFill(start_color=GREY_ALT, end_color=GREY_ALT, fill_type="solid")


def h1_font():
    return Font(name="Calibri", size=12, bold=True, color="FFFFFF")


def body_font():
    return Font(name="Calibri", size=11, color="1F2937")


def sub_font():
    return Font(name="Calibri", size=10, italic=True, color="4B5563")


def _write_header(ws, row, headers, fill_color=GREEN_DARK, height=28):
    for col, text in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=text)
        cell.font = h1_font()
        cell.fill = header_fill(fill_color)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[row].height = height


def _write_row(ws, row, values, is_alt=False):
    for col, val in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = body_font()
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        cell.border = BORDER
        if is_alt:
            cell.fill = alt_fill()


def _set_column_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _add_status_dv(ws, col_letter, first_row, last_row):
    dv = DataValidation(
        type="list",
        formula1='"Not Started,In Progress,Blocked,Done,N/A"',
        allow_blank=True,
        showDropDown=False,
    )
    dv.error = "Choose one of the allowed statuses"
    dv.errorTitle = "Invalid status"
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}{first_row}:{col_letter}{last_row}")


def _add_priority_dv(ws, col_letter, first_row, last_row):
    dv = DataValidation(
        type="list",
        formula1='"Critical,High,Medium,Low"',
        allow_blank=True,
        showDropDown=False,
    )
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}{first_row}:{col_letter}{last_row}")


def _title_block(ws, title, subtitle):
    ws["A1"] = title
    ws["A1"].font = Font(name="Calibri", size=18, bold=True, color=GREEN_DARK)
    ws["A2"] = subtitle
    ws["A2"].font = sub_font()
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 18


# ─── Data ───────────────────────────────────────────────────────────────────

# id, category, platform, task, details, priority, current_status, owner
MASTER = [
    # Accounts
    ("A1", "Accounts", "Apple",   "Apple Developer Program membership",
     "US$99/year. Already enrolled ✓ — verify membership active + payment renewed.",
     "Critical", "Done", "Sanjeev"),
    ("A2", "Accounts", "Google",  "Google Play Console developer account",
     "Organization: CruisE AnalytiX (Account ID 9194497335110928309). Android developer verification DONE ✓.",
     "Critical", "Done", "Sanjeev"),
    ("A3", "Accounts", "Apple",   "App Store Connect app record created",
     "Create app entry, get ASC App ID.", "Critical", "Not Started", "Sanjeev"),
    ("A4", "Accounts", "Google",  "Play Console app record created",
     "WellValet entry exists ✓. Package com.cruiseanalytix.wellvalet. Currently on Internal testing track, status 'Ready to publish'.",
     "Critical", "Done", "Sanjeev"),

    # App identity (already set in app.json)
    ("ID1", "App Identity", "Both", "Bundle ID / package: com.cruiseanalytix.wellvalet",
     "Locked in app.json — do NOT change after first submission.",
     "Critical", "Done", "Prabhat"),
    ("ID2", "App Identity", "Both", "App display name: WellValet",
     "app.json name.", "Critical", "Done", "Prabhat"),
    ("ID3", "App Identity", "Both", "Version 1.0.0 / buildNumber 1",
     "Increment before every submit.", "High", "Done", "Prabhat"),
    ("ID4", "App Identity", "Both", "EAS project linked",
     "projectId 31a587a7-… + owner ecosystem-analytix.",
     "Critical", "Done", "Prabhat"),

    # Legal
    ("L1", "Legal", "Both", "Privacy Policy hosted publicly",
     "Must cover camera, photos, health data (allergies), account, purchase history, AI queries, affiliate partners, analytics. URL: wellvalet.com/privacy.",
     "Critical", "Not Started", "Sanjeev"),
    ("L2", "Legal", "Both", "Terms of Service hosted publicly",
     "URL: wellvalet.com/terms.", "Critical", "Not Started", "Sanjeev"),
    ("L3", "Legal", "Both", "Support email / contact URL",
     "cruise.analytix@gmail.com or dedicated support address.",
     "High", "Done", "Prabhat"),
    ("L4", "Legal", "Apple",  "Privacy Nutrition Labels filled in ASC",
     "Declare every data type collected/linked/tracked.",
     "Critical", "Not Started", "Sanjeev"),
    ("L5", "Legal", "Google", "Data Safety form filled in Play Console",
     "Declare all data types + purpose + encryption in transit.",
     "Critical", "Not Started", "Sanjeev"),
    ("L6", "Legal", "Google", "Content rating (IARC questionnaire)",
     "Complete in Play Console. Health & food related — likely PEGI 3 / Everyone.",
     "High", "Not Started", "Sanjeev"),
    ("L7", "Legal", "Apple",  "Age rating in App Store Connect",
     "4+ or 12+ based on health content questionnaire.",
     "High", "Not Started", "Sanjeev"),
    ("L8", "Legal", "Apple",  "Export compliance",
     "ITSAppUsesNonExemptEncryption = false set in Info.plist.",
     "High", "Done", "Prabhat"),

    # Store listing text
    ("T1", "Store Text", "Both",   "App title (≤30 chars)",
     '"WellValet – Wellness Scanner" (26 chars).',
     "High", "Not Started", "Sanjeev"),
    ("T2", "Store Text", "Apple",  "Subtitle (≤30 chars)",
     "Short tagline. e.g. Scan · Analyse · Improve.",
     "High", "Not Started", "Sanjeev"),
    ("T3", "Store Text", "Google", "Short description (≤80 chars)",
     "Play Store card summary.", "High", "Not Started", "Sanjeev"),
    ("T4", "Store Text", "Both",   "Full description (≤4000 chars)",
     "Features, USP, target audience, CTA.",
     "High", "Not Started", "Sanjeev"),
    ("T5", "Store Text", "Apple",  "Keywords (≤100 chars)",
     "Comma-separated, iOS only.", "Medium", "Not Started", "Sanjeev"),
    ("T6", "Store Text", "Apple",  "Promotional text (≤170 chars)",
     "Can be edited without new build.",
     "Low", "Not Started", "Sanjeev"),
    ("T7", "Store Text", "Both",   "What's New / release notes",
     "For v1.0 — a launch note.", "Medium", "Not Started", "Sanjeev"),

    # Icons & Graphics
    ("G1", "Assets", "Apple",  "iOS App Icon 1024×1024 PNG (no alpha, no rounded corners)",
     "Required for App Store listing.", "Critical", "Not Started", "Designer"),
    ("G2", "Assets", "Google", "Play Store icon 512×512 PNG (32-bit alpha)",
     "Different from adaptive icon.", "Critical", "Not Started", "Designer"),
    ("G3", "Assets", "Google", "Feature graphic 1024×500",
     "Play listing header banner.", "Critical", "Not Started", "Designer"),
    ("G4", "Assets", "Both",   "Adaptive icon (Android) foreground/background/monochrome",
     "Already present in assets/images/.", "High", "Done", "Prabhat"),
    ("G5", "Assets", "Both",   "Splash screen asset production-ready",
     "assets/images/splash-icon.png — verify contrast on dark bg.",
     "High", "Done", "Prabhat"),

    # Screenshots
    ("S1", "Screenshots", "Apple",  "6.7\" iPhone (1290×2796) — min 3, max 10",
     "Mandatory device size for App Store.",
     "Critical", "Not Started", "Designer"),
    ("S2", "Screenshots", "Apple",  "6.5\" iPhone (1284×2778) OR reuse 6.7\"",
     "Optional if only 6.7 provided (Apple upscales).",
     "Medium", "Not Started", "Designer"),
    ("S3", "Screenshots", "Google", "Android phone screenshots (min 1080px shortest edge)",
     "Min 2, max 8.", "Critical", "Not Started", "Designer"),
    ("S4", "Screenshots", "Both",   "Screenshots cover: Scan, Beauty, AI Valet, Meal Planner, Shopping List, Family Plan, Score Ring, Onboarding",
     "8 shots with headline captions.", "High", "Not Started", "Designer"),
    ("S5", "Screenshots", "Apple",  "iPad screenshots",
     "app.json supportsTablet = false — SKIP unless enabling iPad.",
     "N/A", "N/A", "Prabhat"),

    # Native config / build
    ("B1", "Build Config", "Both",   "EAS CLI installed + logged in",
     "npm install -g eas-cli && eas login.",
     "Critical", "Not Started", "Prabhat"),
    ("B2", "Build Config", "Both",   "eas.json production profile",
     "Present ✓ — autoIncrement on iOS.", "High", "Done", "Prabhat"),
    ("B3", "Build Config", "Apple",  "eas.json submit.production.ios values",
     "appleTeamId = KQK3S6DD65 (MIKE SOPHIE) ✓ filled. Still need appleId email + ascAppId.",
     "Critical", "In Progress", "Prabhat"),
    ("B3a", "Build Config", "Apple", "Confirm 'MIKE SOPHIE' Apple Developer account ownership",
     "Team ID KQK3S6DD65 currently registered to MIKE SOPHIE. Confirm this is the account WellValet will launch under, or plan an app transfer.",
     "Critical", "Blocked", "Sanjeev"),
    ("B3b", "Build Config", "Apple", "Apple Distribution certificate exists",
     "Xcode currently shows Development cert only. EAS will auto-generate Distribution cert on first `eas build --profile production` for iOS.",
     "High", "Not Started", "Prabhat"),
    ("B4", "Build Config", "Apple",  "Apple certs + provisioning profiles",
     "Let EAS manage automatically (recommended).",
     "Critical", "Not Started", "Prabhat"),
    ("B5", "Build Config", "Google", "Android keystore backed up",
     "eas credentials → download JKS. Loss = cannot update app forever.",
     "Critical", "Not Started", "Prabhat"),
    ("B6", "Build Config", "Google", "Play service account JSON",
     "Play Console → Setup → API access → create SA → Release Manager role. Save as ./google-play-service-account.json (already gitignored, path wired in eas.json ✓).",
     "Critical", "Not Started", "Prabhat"),
    ("B7", "Build Config", "Both",   "Production build passes locally / on EAS",
     "eas build --platform ios/android --profile production.",
     "Critical", "Not Started", "Prabhat"),

    # Permissions & Info.plist
    ("P1", "Permissions", "Apple",  "NSCameraUsageDescription",
     "Set ✓ — used for barcode + OCR scan.", "Critical", "Done", "Prabhat"),
    ("P2", "Permissions", "Apple",  "NSPhotoLibraryUsageDescription",
     "Set ✓.", "High", "Done", "Prabhat"),
    ("P3", "Permissions", "Apple",  "NSLocationWhenInUseUsageDescription",
     "Set ✓ (Xcode confirmed). Verify if actually used in-app; remove if not.",
     "Medium", "In Progress", "Prabhat"),
    ("P3a", "Permissions", "Apple",  "NSMicrophoneUsageDescription",
     "Visible in Xcode Signing pane. Verify usage — remove if microphone is not used anywhere in code.",
     "Medium", "Not Started", "Prabhat"),
    ("P4", "Permissions", "Google", "Camera + storage permission strings",
     "Auto-generated by Expo — verify manifest.",
     "High", "Not Started", "Prabhat"),

    # In-app purchases / paywall
    ("IP1", "Monetisation", "Apple",  "In-App Purchase products created in ASC",
     "Monthly + Annual + Family SKUs matching backend plan_key.",
     "Critical", "Not Started", "Sanjeev"),
    ("IP2", "Monetisation", "Google", "In-App Products / Subscriptions in Play Console",
     "Same SKUs. Set base plan + offers for trial.",
     "Critical", "Not Started", "Sanjeev"),
    ("IP3", "Monetisation", "Both",   "RevenueCat integration in app",
     "TODO in trial-offer.tsx — replace setPremium mock with RevenueCat.",
     "Critical", "In Progress", "Prabhat"),
    ("IP4", "Monetisation", "Both",   "Tax & Banking info in both consoles",
     "Required before paid apps/IAP go live.",
     "Critical", "Not Started", "Sanjeev"),
    ("IP5", "Monetisation", "Apple",  "Sign in with Apple (if any social login)",
     "Mandatory if Google/other social login exists.",
     "High", "Not Started", "Prabhat"),

    # Testing tracks
    ("TT1", "Testing", "Apple",  "TestFlight internal testing (up to 100)",
     "No beta review. Fastest path.", "High", "Not Started", "Prabhat"),
    ("TT2", "Testing", "Apple",  "TestFlight external testing (up to 10,000)",
     "Requires Beta App Review (~24 hrs).",
     "Medium", "Not Started", "Prabhat"),
    ("TT3", "Testing", "Google", "Play internal testing track",
     "Already active ✓ — build in Internal testing status 'Ready to publish' as of Jul 23 2026. Add testers via Play Console → Testing → Internal testing → Testers.",
     "High", "In Progress", "Prabhat"),
    ("TT4", "Testing", "Google", "Play closed → open testing (staged rollout)",
     "Optional but recommended before Production.",
     "Medium", "Not Started", "Prabhat"),

    # Backend / infra
    ("BE1", "Backend", "Both", "api.wellvalet.com production stable + HTTPS valid",
     "SSL cert valid, rate limiting on, auth flow tested.",
     "Critical", "In Progress", "Backend"),
    ("BE2", "Backend", "Both", "OTA updates channel configured",
     "runtimeVersion 1.0.0 + expo updates URL — already set ✓.",
     "High", "Done", "Prabhat"),
    ("BE3", "Backend", "Both", "Deep linking domain verified",
     "Universal Links (iOS) + App Links (Android) if using https://wellvalet.com/ paths.",
     "Medium", "Not Started", "Prabhat"),
    ("BE4", "Backend", "Both", "Push notifications (if launching with)",
     "APNs cert + FCM sender ID. Currently NOT in scope for v1.0?",
     "Low", "Not Started", "Prabhat"),

    # QA / pre-submit
    ("Q1", "QA", "Both", "Emoji cleanup complete",
     "All emojis replaced with @expo/vector-icons.",
     "High", "Done", "Prabhat"),
    ("Q2", "QA", "Both", "No console.log / debug code in production",
     "Grep for console.log, __DEV__ leaks.",
     "Medium", "Not Started", "Prabhat"),
    ("Q3", "QA", "Both", "Crashlytics / Sentry (optional)",
     "Recommended for prod crash tracking.",
     "Low", "Not Started", "Prabhat"),
    ("Q4", "QA", "Both", "Real-device smoke test — camera, OCR, AI Valet, IAP",
     "Cannot be simulator-only.", "Critical", "Not Started", "Prabhat"),
    ("Q5", "QA", "Both", "Offline behaviour tested",
     "OfflineBanner + graceful fallbacks.",
     "Medium", "Not Started", "Prabhat"),
    ("Q6", "QA", "Both", "Accessibility (VoiceOver + TalkBack sanity)",
     "Icons need accessibilityLabel where they replaced emoji-text.",
     "Medium", "Not Started", "Prabhat"),

    # Submission
    ("SUB1", "Submit", "Apple",  "eas submit --platform ios --latest",
     "Uploads to App Store Connect.", "Critical", "Not Started", "Prabhat"),
    ("SUB2", "Submit", "Google", "eas submit --platform android --latest",
     "Uploads AAB to Play Console track.",
     "Critical", "Not Started", "Prabhat"),
    ("SUB3", "Submit", "Apple",  "Submit for App Review",
     "Answer review notes: test account creds + demo family invite code.",
     "Critical", "Not Started", "Sanjeev"),
    ("SUB4", "Submit", "Google", "Submit to Production / Managed rollout",
     "Start with 10% rollout, monitor crash-free rate.",
     "Critical", "Not Started", "Sanjeev"),
]


APPLE_ONLY = [
    ("Enrolment", "Apple ID with 2FA enabled",
     "Personal Apple ID used for Developer account."),
    ("Enrolment", "D-U-N-S number (Organisation only)",
     "Free from Dun & Bradstreet. 1–5 days verification."),
    ("Enrolment", "Legal entity verification",
     "Ecosystems AnalytiX International Inc. registration docs."),
    ("ASC Setup", "Bank account + Tax forms in Agreements, Tax, and Banking",
     "Mandatory before paid apps or IAP."),
    ("ASC Setup", "Users & Access — invite team members with correct roles",
     "Admin, App Manager, Developer, Marketing."),
    ("ASC Setup", "App-specific password OR ASC API key",
     "For EAS Submit. API key preferred (safer, revocable)."),
    ("App Review", "Demo account credentials for reviewer",
     "Create a permanent test user reviewer can log in with."),
    ("App Review", "App Review notes",
     "Explain barcode scanning, allergen alerts, affiliate links."),
    ("App Review", "Contact info (phone + email)",
     "Reviewer must be able to reach you within 48h."),
    ("App Review", "Content rights",
     "Confirm all product images, brand logos have rights or use public data (OpenFoodFacts etc)."),
    ("Guidelines", "4.2 Minimum Functionality",
     "Must be more than a webview. WellValet has native camera + logic ✓."),
    ("Guidelines", "5.1 Privacy — Data collection consent",
     "Login/signup must show TOS + Privacy links."),
    ("Guidelines", "5.1.1(v) Account deletion in-app",
     "MANDATORY. Add 'Delete Account' in profile — verify implemented."),
    ("Guidelines", "3.1.1 IAP for digital goods",
     "Premium unlock MUST use IAP, not external Stripe/URL."),
    ("Guidelines", "1.4.1 Physical harm — Health claims",
     "Add disclaimer: 'Not medical advice'. Present in disclaimers ✓."),
]


PLAY_ONLY = [
    ("Enrolment", "Google account with 2FA enabled",
     "Business email preferred."),
    ("Enrolment", "Developer verification (ID + address)",
     "New Play requirement since 2023."),
    ("Enrolment", "Organisation D-U-N-S if publishing as company",
     "Required for company account since Sep 2023."),
    ("Console Setup", "Merchant account (Google Payments)",
     "Required for paid apps / IAP."),
    ("Console Setup", "API access → Service account with Release Manager role",
     "JSON key used by EAS Submit."),
    ("Console Setup", "App signing by Google Play (default, recommended)",
     "Upload key vs. app signing key — EAS handles upload key."),
    ("Policies", "Target API level ≥ 34 (Android 14) by Aug 2025",
     "Expo SDK 51+ compiles against 34 automatically."),
    ("Policies", "App bundle (.aab) only — no APK for Production",
     "eas.json production profile builds AAB by default ✓."),
    ("Policies", "64-bit only requirement",
     "Expo/RN handles by default."),
    ("Policies", "Health apps declaration",
     "Play Console asks about medical/wellness claims — answer honestly."),
    ("Policies", "News apps declaration",
     "No."),
    ("Policies", "COVID-19 apps declaration",
     "No."),
    ("Policies", "Financial features",
     "No — until IAP live."),
    ("Policies", "Family policy compliance",
     "If targeting under-13 → additional constraints. Target 13+."),
    ("Store Listing", "App category: Health & Fitness (primary), Food & Drink (secondary)",
     "Cannot pick 2 primaries."),
    ("Store Listing", "Tags (up to 5)",
     "Pick from Play's tag catalogue."),
    ("Store Listing", "Contact details (email, phone, website)",
     "Website: wellvalet.com."),
    ("Store Listing", "External marketing opt-in",
     "Play marketing surfaces."),
]


ASSETS = [
    ("iOS App Icon",         "1024×1024",     "PNG, no alpha, sRGB", "Apple",  "App Store listing"),
    ("iOS Adaptive Icons",   "Multiple sizes","Handled by Expo",     "Apple",  "Auto-generated from icon"),
    ("Play Store Icon",      "512×512",       "PNG, 32-bit alpha",   "Google", "Play listing"),
    ("Feature Graphic",      "1024×500",      "PNG or JPG",          "Google", "Play listing header"),
    ("Adaptive Icon Fg",     "108×108 dp / 432×432 px", "PNG",       "Google", "app.json ✓"),
    ("Adaptive Icon Bg",     "108×108 dp",    "PNG or solid color",  "Google", "app.json ✓"),
    ("Monochrome Icon",      "108×108 dp",    "PNG",                 "Google", "Themed icon (A13+)"),
    ("iPhone 6.7\" Shot",    "1290×2796",     "PNG or JPG",          "Apple",  "Mandatory size"),
    ("iPhone 6.5\" Shot",    "1284×2778",     "PNG or JPG",          "Apple",  "Optional if 6.7 provided"),
    ("iPhone 5.5\" Shot",    "1242×2208",     "PNG or JPG",          "Apple",  "Optional (older devices)"),
    ("iPad 12.9\" Shot",     "2048×2732",     "PNG or JPG",          "Apple",  "SKIP (iPad off)"),
    ("Android Phone Shot",   "≥1080 shortest side, 16:9 or 9:16", "PNG or JPG", "Google", "Min 2, max 8"),
    ("Android Tablet Shot",  "7\" + 10\"",     "PNG or JPG",          "Google", "Optional"),
    ("App Preview Video",    "Vertical H.264, ≤30s", "MOV/MP4",       "Apple",  "Optional but boosts conversion"),
    ("Promo Video",          "YouTube URL",   "Any format on YT",    "Google", "Optional"),
    ("Splash Screen",        "1284×2778 base","PNG",                 "Both",   "app.json splash asset ✓"),
]


LEGAL = [
    ("Privacy Policy", "Both", "Public URL required. Must list ALL data collected + purpose + third parties.",
     "Camera images (not stored server-side), account data, wellness profile (allergies — sensitive), purchase history, AI Valet queries, affiliate click IDs.", "Not Started"),
    ("Terms of Service", "Both", "Public URL. Cover subscription auto-renew, cancellation, refund policy, liability limitations.",
     "Include disclaimer: 'WellValet is not medical advice.'", "Not Started"),
    ("Data Safety Form", "Google", "Play Console → App content → Data safety.",
     "Declare: Personal info (email, name), Health & fitness (allergies), Purchase history, App activity (scans), Camera. Encrypted in transit = YES.", "Not Started"),
    ("Privacy Nutrition Labels", "Apple", "ASC → App Privacy → Manage.",
     "Data types linked to user: Contact Info, Health & Fitness, Purchases, User Content, Identifiers. Purposes: App Functionality, Analytics.", "Not Started"),
    ("Age Rating (Apple)", "Apple", "Answer questionnaire in ASC.",
     "Health/Medical treatment info → Infrequent/Mild → Rated 4+ or 12+.", "Not Started"),
    ("Content Rating (Google)", "Google", "IARC questionnaire.",
     "Likely PEGI 3 / Everyone.", "Not Started"),
    ("Health disclaimer visible in-app", "Both", "Show 'Not medical advice' near scan results.",
     "Already present in scan disclaimers ✓.", "Done"),
    ("Account deletion in-app", "Both", "Apple Guideline 5.1.1(v) MANDATORY.",
     "Verify profile screen has Delete Account button + backend endpoint.", "In Progress"),
    ("Cookie / tracking consent", "Both", "If any analytics SDK (Firebase, Amplitude, etc.) is added later.",
     "Currently none integrated.", "N/A"),
    ("EULA (optional)", "Apple", "Custom EULA or default Apple EULA.",
     "Default works unless custom clauses needed.", "N/A"),
    ("GDPR readiness", "Both", "If EU users targeted — right to erasure, data export.",
     "Backend endpoints for /account/delete + /account/export.", "Not Started"),
    ("HIPAA scope check", "Both", "US medical data — NOT HIPAA regulated since not a covered entity.",
     "Wellness app, not medical provider — HIPAA not applicable.", "N/A"),
]


# ─── Sheet builders ─────────────────────────────────────────────────────────

def build_master(wb):
    ws = wb.create_sheet("Master Checklist")
    _title_block(ws,
                 "WellValet — Store Launch Checklist",
                 "Complete every Critical + High item before submitting to review. Sanjeev = business/legal, Prabhat = engineering.")

    header_row = 4
    _write_header(ws, header_row, [
        "ID", "Category", "Platform", "Task", "Details / Notes",
        "Priority", "Status", "Owner",
    ])

    for i, row in enumerate(MASTER, start=header_row + 1):
        _write_row(ws, i, row, is_alt=(i % 2 == 0))
        # Colour priority cell
        p = ws.cell(row=i, column=6)
        if p.value == "Critical":
            p.fill = PatternFill(start_color="FFE4E6", end_color="FFE4E6", fill_type="solid")
            p.font = Font(bold=True, color=RED)
        elif p.value == "High":
            p.fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
            p.font = Font(bold=True, color="B45309")
        # Colour status
        s = ws.cell(row=i, column=7)
        if s.value == "Done":
            s.fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
            s.font = Font(bold=True, color=GREEN_DARK)
        elif s.value == "In Progress":
            s.fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
            s.font = Font(bold=True, color=BLUE)
        elif s.value == "Blocked":
            s.fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
            s.font = Font(bold=True, color=RED)
        elif s.value == "N/A":
            s.fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
            s.font = Font(italic=True, color="6B7280")

    last = header_row + len(MASTER)
    _add_priority_dv(ws, "F", header_row + 1, last)
    _add_status_dv(ws, "G", header_row + 1, last)

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A{header_row}:H{last}"
    _set_column_widths(ws, [8, 16, 10, 40, 60, 12, 14, 14])


def build_platform(wb, name, rows, header_color):
    ws = wb.create_sheet(name)
    _title_block(ws,
                 f"{name} — Specific Requirements",
                 "Complements the Master Checklist with store-specific policy & setup items.")

    header_row = 4
    _write_header(ws, header_row, ["Section", "Item", "Details"], fill_color=header_color)
    for i, row in enumerate(rows, start=header_row + 1):
        _write_row(ws, i, row, is_alt=(i % 2 == 0))

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A{header_row}:C{header_row + len(rows)}"
    _set_column_widths(ws, [18, 45, 70])


def build_assets(wb):
    ws = wb.create_sheet("Assets Reference")
    _title_block(ws, "Asset Sizes & Formats", "Everything the designer must ship.")

    header_row = 4
    _write_header(ws, header_row,
                  ["Asset", "Size", "Format", "Platform", "Notes"],
                  fill_color=ORANGE)
    for i, row in enumerate(ASSETS, start=header_row + 1):
        _write_row(ws, i, row, is_alt=(i % 2 == 0))

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A{header_row}:E{header_row + len(ASSETS)}"
    _set_column_widths(ws, [26, 32, 26, 12, 45])


def build_legal(wb):
    ws = wb.create_sheet("Legal & Compliance")
    _title_block(ws, "Legal, Privacy & Compliance",
                 "Store rejections most often happen here. Get these right.")

    header_row = 4
    _write_header(ws, header_row,
                  ["Item", "Platform", "Requirement", "What to declare / do", "Status"],
                  fill_color=RED)
    for i, row in enumerate(LEGAL, start=header_row + 1):
        _write_row(ws, i, row, is_alt=(i % 2 == 0))
        s = ws.cell(row=i, column=5)
        if s.value == "Done":
            s.fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
            s.font = Font(bold=True, color=GREEN_DARK)
        elif s.value == "In Progress":
            s.fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
            s.font = Font(bold=True, color=BLUE)
        elif s.value == "N/A":
            s.fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
            s.font = Font(italic=True, color="6B7280")

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A{header_row}:E{header_row + len(LEGAL)}"
    _set_column_widths(ws, [26, 12, 45, 60, 16])


def build_summary(wb):
    ws = wb.create_sheet("Summary", 0)  # first sheet
    _title_block(ws, "WellValet — Launch Readiness Summary",
                 "Auto-generated on: run date. Refer to Master Checklist for authoritative status.")

    # Counters
    cats = {}
    prios = {"Critical": 0, "High": 0, "Medium": 0, "Low": 0}
    statuses = {"Not Started": 0, "In Progress": 0, "Blocked": 0, "Done": 0, "N/A": 0}
    owners = {}
    for _id, cat, plat, task, det, pri, stat, owner in MASTER:
        cats[cat] = cats.get(cat, 0) + 1
        prios[pri] = prios.get(pri, 0) + 1
        statuses[stat] = statuses.get(stat, 0) + 1
        owners[owner] = owners.get(owner, 0) + 1

    def _section(title, row):
        c = ws.cell(row=row, column=1, value=title)
        c.font = Font(size=13, bold=True, color=GREEN_DARK)
        return row + 1

    r = 4
    r = _section("By Priority", r)
    _write_header(ws, r, ["Priority", "Count"], height=22)
    r += 1
    for k in ["Critical", "High", "Medium", "Low"]:
        _write_row(ws, r, [k, prios.get(k, 0)], is_alt=(r % 2 == 0))
        r += 1

    r += 1
    r = _section("By Status", r)
    _write_header(ws, r, ["Status", "Count"], height=22)
    r += 1
    for k in ["Not Started", "In Progress", "Blocked", "Done", "N/A"]:
        _write_row(ws, r, [k, statuses.get(k, 0)], is_alt=(r % 2 == 0))
        r += 1

    r += 1
    r = _section("By Category", r)
    _write_header(ws, r, ["Category", "Count"], height=22)
    r += 1
    for k in sorted(cats.keys()):
        _write_row(ws, r, [k, cats[k]], is_alt=(r % 2 == 0))
        r += 1

    r += 1
    r = _section("By Owner", r)
    _write_header(ws, r, ["Owner", "Count"], height=22)
    r += 1
    for k in sorted(owners.keys()):
        _write_row(ws, r, [k, owners[k]], is_alt=(r % 2 == 0))
        r += 1

    r += 2
    ws.cell(row=r, column=1,
            value="Recommended launch order").font = Font(size=13, bold=True, color=GREEN_DARK)
    r += 1
    steps = [
        "1. Open Apple Developer + Play Console accounts (Sanjeev)",
        "2. Draft Privacy Policy + Terms and host at wellvalet.com (Sanjeev)",
        "3. Fill eas.json submit config with real Apple IDs (Prabhat)",
        "4. Designer ships icon 1024, feature graphic, 8 screenshots (Designer)",
        "5. RevenueCat + IAP wiring in trial-offer.tsx (Prabhat)",
        "6. First EAS production build for iOS + Android (Prabhat)",
        "7. TestFlight internal + Play internal test with 5–10 people (Prabhat + team)",
        "8. Fix crash/UX issues found in beta",
        "9. Fill store listings, screenshots, data safety, nutrition labels (Sanjeev)",
        "10. Submit for Apple review + Play production (staged 10% rollout)",
    ]
    for step in steps:
        c = ws.cell(row=r, column=1, value=step)
        c.font = body_font()
        c.alignment = Alignment(wrap_text=True, vertical="top")
        r += 1

    _set_column_widths(ws, [90, 14])


# ─── Main ───────────────────────────────────────────────────────────────────

def main():
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    build_master(wb)
    build_platform(wb, "Apple App Store", APPLE_ONLY, "1D4ED8")
    build_platform(wb, "Google Play",     PLAY_ONLY,  "047857")
    build_assets(wb)
    build_legal(wb)
    build_summary(wb)  # inserted at index 0

    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
